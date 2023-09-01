from django.core.management.base import BaseCommand
import openpyxl
from college.models import *


def create_or_update_gpa(student, gpa):
    gpa_point = Point.objects.filter(student=student, point_type=0).first()
    if gpa_point:
        gpa_point.value = gpa
    else:
        gpa_point = Point(student=student, point_type=0, value=gpa)

    gpa_point.save()


def insert_point(student, semester_idx, value):
    p = Point.objects.filter(student=student, point_type=semester_idx).first()
    if p:
        return

    new_point = Point(student=student, point_type=semester_idx, value=value)
    new_point.save()


class Command(BaseCommand):
    def handle(self, *args, **options):
        MAX_ROW = 4058
        COLS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']

        wb = openpyxl.load_workbook("points.xlsx")
        ws = wb.active

        for r in range(1, MAX_ROW):
            mssv = ws[f"{COLS[0]}{r}"].value
            gpa = ws[f"{COLS[1]}{r}"].value
            s1 = ws[f"{COLS[2]}{r}"].value
            s2 = ws[f"{COLS[3]}{r}"].value
            s3 = ws[f"{COLS[4]}{r}"].value
            s4 = ws[f"{COLS[5]}{r}"].value
            s5 = ws[f"{COLS[6]}{r}"].value
            s6 = ws[f"{COLS[7]}{r}"].value
            s7 = ws[f"{COLS[8]}{r}"].value
            s8 = ws[f"{COLS[9]}{r}"].value
            s9 = ws[f"{COLS[10]}{r}"].value
            s10 = ws[f"{COLS[11]}{r}"].value

            gpa = gpa if gpa is not None else 0
            s1 = s1 if s1 is not None else 0
            s2 = s2 if s2 is not None else 0
            s3 = s3 if s3 is not None else 0
            s4 = s4 if s4 is not None else 0
            s5 = s5 if s5 is not None else 0
            s6 = s6 if s6 is not None else 0
            s7 = s7 if s7 is not None else 0
            s8 = s8 if s8 is not None else 0
            s9 = s9 if s9 is not None else 0
            s10 = s10 if s10 is not None else 0

            try:
                student = Student.objects.get(pk=mssv)
                if not student:
                    print(f"Kiểm tra: {mssv}")

                if gpa > 0:
                    create_or_update_gpa(student, gpa)

                if s1 > 0:
                    insert_point(student, 1, s1)

                if s2 > 0:
                    insert_point(student, 2, s2)

                if s3 > 0:
                    insert_point(student, 3, s3)

                if s4 > 0:
                    insert_point(student, 4, s4)

                if s5 > 0:
                    insert_point(student, 5, s5)

                if s6 > 0:
                    insert_point(student, 6, s6)

                if s8 > 0:
                    insert_point(student, 7, s7)

                if s8 > 0:
                    insert_point(student, 8, s8)

                if s9 > 0:
                    insert_point(student, 9, s9)

                if s10 > 0:
                    insert_point(student, 10, s10)

            except Exception as e:
                print(e)
